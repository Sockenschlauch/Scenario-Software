import openpyxl
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side


# defining the style of borders
thin_border_bottom = Border(bottom=Side(style='thin'))
thin_border_right = Border(right=Side(style='thin'))


def read_projections(path):  # reads the list of factors with projections
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    cell_value = sheet
    factor_list = []

    i = 0
    for row in sheet.iter_rows():
        factor_list.append([])
        for cell in row:
            if cell.value:
                factor_list[i].append(cell.value)
        i += 1
    return factor_list


# function for printing the projections in the terminal for debugging
def print_projections(factor_list):
    for factor in factor_list[1:]:
        print(factor[0], end=": ")
        for projection in factor[1:-1]:
            print(projection, end=", ")
        print(factor[-1])


def write_consistency_matrix(path, factor_list):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "consistency matrix"

    len = get_projection_count(factor_list)

# Writing the vertical projections
    row = 3
    colum = 1
    for factor in factor_list[1:]:
        for projection in factor[1:]:
            sheet.cell(row, colum).value = factor[0]
            sheet.cell(row, colum+1).value = projection
            sheet.cell(row, colum+1).border = thin_border_right
            row += 1

        for i in range(1, len+3):
            if sheet.cell(row-1, i).border == Border(right=Side(style='thin')):
                sheet.cell(row-1, i).border = Border(bottom=Side(style='thin'),
                                                     right=Side(style='thin'))
            else:
                sheet.cell(row-1, i).border = thin_border_bottom

# Writing the horizontal projections
    row = 1
    colum = 3
    for factor in factor_list[1:]:
        for projection in factor[1:]:
            sheet.cell(row, colum).value = factor[0]
            sheet.cell(row+1, colum).value = projection
            sheet.cell(row+1, colum).border = thin_border_bottom
            colum += 1

        for i in range(1, len+3):
            if sheet.cell(i, colum-1).border == Border(bottom=Side(style='thin')):
                sheet.cell(i, colum-1).border = Border(bottom=Side(style='thin'),
                                                       right=Side(style='thin'))
            else:
                sheet.cell(i, colum-1).border = thin_border_right

    wb.save(filename=path)


def get_projection_count(factor_list):  # returns the length of the factor list
    factor_list_len = 0
    for factor in factor_list[1:]:
        for projection in factor[1:]:
            factor_list_len += 1
    return factor_list_len
