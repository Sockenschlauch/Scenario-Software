import openpyxl
from openpyxl import Workbook


def write_clusters(factors, projections, cluster_list):
    dst_filename = "clusters.xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.title = "consistency matrix"

    max = 0
    # Getting the maximum number of projections per row
    for factor in projections:
        if len(factor) > max:
            max = len(factor)

    # Sorting the cluster_list
    cluster_list.sort(key=lambda x: x.cluster)

    # Getting the mixture list
    mixtures = []
    for cluster in cluster_list:
        mixtures.append(cluster.get_mixture())

    # Writing the factors and projections
    for i in range(len(factors)):
        sheet.cell((i+1)*2, 1).value = factors[i]
        for j in range(len(cluster_list)):
            for k in range(len(projections[i])):
                sheet.cell((i+1)*2, 2+k+max*j).value = projections[i][k]
                sheet.cell((i+1)*2+1, 2+k+max*j).value = mixtures[j][i][k]

    # Writing the cluster names
    for i, cluster in enumerate(cluster_list):
        for j in range(max):
            sheet.cell(1, 2+j+max*i).value = "cluster " + str(cluster.cluster)

    # Writing to the file
    wb.save(filename=dst_filename)
