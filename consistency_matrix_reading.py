import openpyxl
from openpyxl import Workbook


# reads the projections from consistency matrix
def read_projections_from_matrix(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    cell_value = sheet

    row = 3
    colum = 1

    projection_list = [[]]
    factor = sheet.cell(row, colum).value

    while sheet.cell(row, colum).value:
        if sheet.cell(row, colum).value == factor:
            projection_list[-1].append(sheet.cell(row, colum+1).value)
        else:
            projection_list.append([])
            projection_list[-1].append(sheet.cell(row, colum+1).value)
            factor = sheet.cell(row, colum).value
        row += 1

    return projection_list


# reads the factors from consistency matrix
def read_factors_from_matrix(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    cell_value = sheet

    row = 3
    colum = 1

    factor_list = [sheet.cell(row, colum).value]

    while sheet.cell(row, colum).value:
        if sheet.cell(row, colum).value is not factor_list[-1]:
            factor_list.append(sheet.cell(row, colum).value)
        else:
            row += 1

    return factor_list


def read_matrix(path, projections):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    cell_value = sheet

    matrix = []
    row = 3
    colum = 3
    for horizontal in range(len(projections)):
        matrix.append([])
        for vertical in range(len(projections)):
            matrix[horizontal].append([])
            if horizontal < vertical:
                for x in range(len(projections[horizontal])):
                    matrix[horizontal][vertical].append([])
                    for y in range(len(projections[vertical])):
                        matrix[horizontal][vertical][x].append(sheet.cell(
                            row + y, colum + x).value)
            row += len(projections[vertical])
        colum += len(projections[horizontal])
        row = 3
    return matrix
