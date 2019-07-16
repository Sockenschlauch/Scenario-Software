import openpyxl
from openpyxl import Workbook

path = r"E:\Git\Scenario-Software\SimpleTest_Factors.xlsx"


def read_projections(path):  # reads the list of factors with projections
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    cell_value = sheet
    factor_list = []

    i = 0
    for row in sheet.iter_rows():
        factor_list.append([])
        for cell in row:
            if cell.value:
                factor_list[i].append(cell.value)
        i += 1
    return factor_list


# function for printing the projections in the terminal for debugging
def print_projections(factor_list):
    for factor in factor_list[1:]:
        print(factor[0], end=": ")
        for projection in factor[1:-1]:
            print(projection, end=", ")
        print(factor[-1])


def write_consistency_matrix(path, factor_list):
    dst_filename = "consistency_matrix.xlsx"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "consistency matrix"

# Writing the vertical projections
    row = 3
    colum = 1
    for factor in factor_list[1:]:
        for projection in factor[1:]:

            ws1.cell(row, colum).value = factor[0]
            ws1.cell(row, colum+1).value = projection

            row += 1
# Writing the horizontal projections
    row = 1
    colum = 3
    for factor in factor_list[1:]:
        for projection in factor[1:]:

            ws1.cell(row, colum).value = factor[0]
            ws1.cell(row+1, colum).value = projection

            colum += 1

    wb.save(filename=dst_filename)


def main():
    projections = read_projections(path)
    print("Creating consistency matrix...")
    write_consistency_matrix(path=path, factor_list=projections)
    print("Consistency matrix created!")


if __name__ == "__main__":
    main()
